from openpyxl import Workbook, load_workbook
import re
import os


def prochain_mot(pos,chaine):#trouve le prochain mot, strictement après la position pos dans la chaine chaine
        interruption=False
        mot=""
        i=pos+1
        l=len(chaine)
        if chaine[i]=="{":#On ne tient pas compte des déterminatifs
                determinant=True
                while determinant:
                        i+=1
                        if chaine[i]=="}":
                                determinant=False
                                i+=1
        if chaine[i]=="_":#pour les sumérogrammes. On élimine simplement le signe indiquant que c'est un sumérogramme.
                i+=1
        while chaine[i]==" ":#élimine les espaces devant
                i+=1
        while (not interruption) and (i<l):
                if chaine[i] == " ":
                        interruption=True
                        continue
                mot=mot+chaine[i]
                i+=1
        return(mot)

def determinant_avant(pos, chaine):#trouve le déterminatif qui précède la position pos dans la chaine chaine.
        i=pos
        if not (chaine[i]=="}"):
                cont=True
                while cont:
                        i-=1
                        if i<0:
                                return("")
                        if (chaine[i]==" "):#Dans ce cas, on ne renvoie rien, car le mot qui contient la position pos n'admet pas de déterminatif le précédant.
                                return("")
                        if chaine[i]== "}":
                                cont=False
                #return("")
        det=""
        while (not chaine[i]=="{") and (i>-1):
                i-=1
                det=chaine[i]+det
        return(det[1:])

def determinant_apres(pos,chaine):#retourne le déterminatif qui succède au mot situé à la position pos.
        i=pos
        l=len(chaine)
        if not (chaine[i]=="{"):
                cont=True
                while cont:
                        i+=1
                        if i>l-1:
                                return("")
                        if (chaine[i]==" "):#Car cela veut dire qu'il n'y a pas de déterminatif accolé à la fin du mot considéré.
                                return("")
                        if chaine[i]=="{":
                                cont=False
                #return("")
        det=""

        while (not chaine[i]=="}") and (i<l):
                i+=1
                det=det+chaine[i]
        return(det)


#Système S
systeme_S={"disz":1,"u":10,"gesz2":60,"geszu":600,"szar2":3600}

#Analyse ce qui précède l'unité (sachant que pos se situe juste avant le début de l'unité) dans la ligne chaine. Le but étant de trouver les nombres précédant l'unité. 
def analyse_avant(pos,chaine):
        fini=False#Permet de savoir si on a fini de parser tous les nombres précédant l'unité.
        position_courante=pos
        def isoler_mot(p,c):#isole le mot qui précède; p est une position et c une chaîne de caractère.
                interruption=False
                i=p
                mot=""
                if c[i]=="}":
                        determinant=True
                        while determinant:
                                i-=1
                                if c[i]=="{":
                                        determinant=False
                                        i-=1
                if c[i]=="_":
                        i-=1
                while c[i]==" ":
                        i-=1
                while (not interruption) and (i>-1):
                        if c[i]== " ":
                                interruption=True
                                continue
                        mot=c[i]+mot
                        i-=1
                return(i,mot)
        total=float(0)
        while (not fini) and position_courante>-1:
                position_courante,mot_courant=isoler_mot(position_courante,chaine)
                for cle,valeur in systeme_S.items():
                        m=re.search(r"(\S*)\("+cle+r"\)\W*",mot_courant) #On cherche si des mots-clefs du système S
                        if m:
                                nombre=m.group(1)
                                fraction=re.search(r"(\D*)([0-9]*)/([0-9]*)",nombre)
                                if fraction:
                                        total+=float(fraction.group(2))/float(fraction.group(3))*valeur
                                else:
                                        entier=re.search(r"(\D*)([0-9]*)",nombre)
                                        if entier:
                                                total+=float(entier.group(2))*valeur
                        else:
                                fini=True
        return(total)
                





#Et pour regarder un mot qui précède strictement pos dans la chaine chaine:
def mot_davant(pos,chaine):
        #ce qui suit permet de trouver le mot juste précédant
        interruption=False
        mot=""
        i=pos
        #On élimine les déterminants, etc.:
        if chaine[i]=="}":
                determinant=True
                while determinant:
                        i-=1
                        if chaine[i]=="{":
                                determinant=False
                                i-=1
        if chaine[i]=="_":#ce caractère indique que l'on a une partie en sumérien
                i-=1
        while chaine[i]==" ":
                i-=1
        while (not interruption) and (i>-1):
                if chaine[i]==" ":
                        interruption=True
                        continue
                mot=chaine[i]+mot
                i-=1
        return(mot)

def beginning_recipe(ligne):#Controle si on est au début d'un nouveau texte
        if ligne==[]:
                return(False)
        if ligne[0]=="&":
                return(True)
        return(False)



#Fonction pour voir si une nouvelle recette commence dans la ligne l. On revoit toutes les sous-listes, ainsi que le nombre d'occurences.
def nouvelle_recette(l):
        debut_recette=r"disz (?:na|kimin|ki\.min|ki-min|lu2)"
        a=re.split(debut_recette,l.lower())
        return(len(a)-1,a)






workbook=Workbook()
sheet = workbook.active #First sheet
sheet.title="Raw data"

#Pour donner des informations sur l'encodage
sh_metadata = workbook.create_sheet("CDLI Metadata",1)




Unites_de_mesure=["ban2","gin2","ma-na","sila3","barig","gur","sze","gu2","gun2"]
Nombre_unites=[0 for i in range(len(Unites_de_mesure))]#nombre de fois où l'unité est employé dans une recette


#Pour analyser les mesures plus précisément
sh_gin=[]
for i in range(len(Unites_de_mesure)):
        sh_gin.append(workbook.create_sheet("Utilisation "+Unites_de_mesure[i],2+i))
        sh_gin[i]["A1"].value="Texte"
        sh_gin[i]["B1"].value="Rang dans le texte"
        sh_gin[i]["C1"].value="Mot précédent l'unité"
        sh_gin[i]["F1"].value="Mot suivant l'unité"
        sh_gin[i]["D1"].value="Déterminatif précédant l'unité"
        sh_gin[i]["E1"].value="Déterminatif suivant l'unité"
        sh_gin[i]["G1"].value="Rang de la recette dans le texte"
        sh_gin[i]["H1"].value="Quantité"

#Lecture du fichier cdli texte
chemin_data=os.path.join(os.path.dirname(os.path.abspath(__file__)),"Data")#chemin vers les données.
fichier_cdli=open(os.path.join(chemin_data,"cdli_subgenre-medical_allperiod.txt"),"r")
Liste_lignes_cdli=fichier_cdli.readlines()
fichier_cdli.close()

#Pour compter le numéro du texte et savoir où on en est dans le fichier Excel
numero_texte=0

#Première ligne
sheet["A1"].value="Numéro de texte"
dd=len(Unites_de_mesure)
for i in range(dd):
        sheet.cell(row=1,column=i+2).value=Unites_de_mesure[i]
sheet.cell(row=1,column=dd+2).value="Traduction?"#pour savoir si une traduction est disponible
sheet.cell(row=1,column=dd+3).value="Langue"
sheet.cell(row=1,column=dd+4).value="Epoque"
sheet.cell(row=1,column=dd+5).value="Provenance"
sheet.cell(row=1,column=dd+6).value="Nombre de recettes probablement contenues"#évalue le nombre de recettes contenues grâce à la fonction nouvelle_recette



#fonction d'écriture des données d'une ligne dans Excel
def ecrire_ligne(numero_ligne,Nb_unites, est_trad,language,nb_rec):
        d=len(Unites_de_mesure)
        for i in range(d):
                sheet.cell(row=numero_ligne,column=i+2).value=Nb_unites[i]
        if est_trad:
                sheet.cell(row=numero_ligne,column=d+2).value= "X"
        if not language=="":
                sheet.cell(row=numero_ligne,column=d+3).value=language
        sheet.cell(row=numero_ligne,column=d+6).value=nb_rec


#Pour regarder les mots précédant ou suivant les unités de mesure
#C'est une liste telle que Mots_precedants_suivants[i] contienne une liste de toutes les occurences de la mesure i, et donc Mots_precedants_suivants[i][j] est une liste à 7 éléments décrivant la jème occurence de l'unité i
#Cette liste à 7 éléments contient, dans l'ordre: la référence du texte, le mot qui précède, le déterminant qui rpécède, le déterminant qui succède, le mot qui succède, le numéro de recette
Mots_precedants_suivants=[[] for i in range(len(Unites_de_mesure))]

#Pour stocker les quantités
#Quantites_unites=[[] for i in range(len(Unites_de_mesure))]


est_traduit=False # dit si le texte dispose d'une traduction
langue=""#langue si indiquée


#Pour voir les langues différentes référencées
Liste_langues=[]

for ligne_texte in Liste_lignes_cdli:
        if beginning_recipe(ligne_texte):
                nom=ligne_texte[1:]#Référence du texte
                
                #On finit d'écrire ce qui est relatif au texte précédent
                if numero_texte>0:
#                        for i in range(len(Unites_de_mesure)):
#                                sheet.cell(row=numero_texte+1,column=i+2).value=Nombre_unites[i]
                        ecrire_ligne(numero_texte+1, Nombre_unites,est_traduit, langue,nombre_recette_dans_texte)

                nombre_recette_dans_texte=0#on écrit le nombre de recettes contenues dans ce texte
                numero_texte+=1
                sheet["A"+str(numero_texte+1)].value = nom#On met le nom du nouveau texte
                Nombre_unites=[0 for i in range(len(Unites_de_mesure))]#on remet à zéro
                est_traduit=False
                langue=""
                if numero_texte%10==0:
                        print(numero_texte)#Permet simplement de se rendre compte que le programme travaille bien quand on l'execute.

        #On regarde ce qui se passe pour des lignes particulières
        if ligne_texte[0]=="#":
        #Si la tablette est traduite
                if not re.search("#tr.en|#tr-en",ligne_texte.lower())==None:
                        est_traduit=True
                        continue

        #La langue
                if not re.search("#atf: lang",ligne_texte.lower())==None:
                        langue=prochain_mot(9,ligne_texte)
                        if not langue in Liste_langues:
                                Liste_langues.append(langue)

        #On regarde si une nouvelle recette commence:
        nbr,liste_sous_lignes=nouvelle_recette(ligne_texte)
        nombre_recette_dans_texte+=nbr

        for riga in liste_sous_lignes:
                #on itère sur la ligne pour y trouver les ingrédients
                for i in range(len(Unites_de_mesure)):
                        for match in re.finditer("( |_|\})"+Unites_de_mesure[i]+"( |_|#|\{|\?|\*|\!)",riga.lower()):
                                Nombre_unites[i]+=1
                                mot_suivant=prochain_mot(match.end()-1,riga)
                                det_precedant=determinant_avant(match.end()-1,riga)
                                det_suivant=determinant_apres(match.start(),riga)
                                mot_precedant=mot_davant(match.start(),riga)
                                q=analyse_avant(match.start(),riga)
                                Mots_precedants_suivants[i].append([nom,mot_precedant,det_precedant,det_suivant,mot_suivant,nombre_recette_dans_texte,q])
                                                        
        
        

#On s'assure que la dernière ligne soit écrite:
ecrire_ligne(numero_texte+1, Nombre_unites,est_traduit,langue,nombre_recette_dans_texte)
##for i in range(len(Unites_de_mesure)):
##        sheet.cell(row=numero_texte+1,column=i+2).value=Nombre_unites[i]






#On veut regarder tous les mots clefs disponibles après #
sh_metadata["A1"].value="Informations données par #"

sh_metadata["B1"].value="Numéro du premier texte d'apparition"
Liste_possibles=[]
i=0#pour la ligne d'Excel
j=0#pour le numéro de texte
for l in Liste_lignes_cdli:
        if l==[]:
                continue
        if beginning_recipe(l):
                j+=1
        if l[0]=="#":
                m=prochain_mot(0,l)#premier mot après #
                if not m in Liste_possibles:
                        Liste_possibles.append(m)
                        i+=1
                        sh_metadata["A"+str(i+1)].value=m
                        sh_metadata["B"+str(i+1)].value=j
                        print(i)#histoire de savoir où on en est
        
#On ajoute les langues
sh_metadata["A"+str(i+3)].value="Langues différentes référencées"
for k in range(len(Liste_langues)):
        sh_metadata["A"+str(i+4+k)].value=Liste_langues[k]
        print(k)#Pour savoir où on en est


#On rajoute les mots suivant, précédant, etc gin2

for k in range(len(Unites_de_mesure)):
        i=0#pour compter les lignes de la worksheet
        j=0#pour compter les occurences dans un texte
        premier_texte=Mots_precedants_suivants[1][0][0]
        for occurence in Mots_precedants_suivants[k]:
                i+=1
                j+=1
                sh_gin[k]["A"+str(i+2)].value=occurence[0]
                if not (occurence[0]==premier_texte):
                        premier_texte=occurence[0]
                        j=1
                sh_gin[k]["B"+str(i+2)].value=j
                sh_gin[k]["C"+str(i+2)].value=occurence[1]
                sh_gin[k]["D"+str(i+2)].value=occurence[2]
                sh_gin[k]["E"+str(i+2)].value=occurence[3]
                sh_gin[k]["F"+str(i+2)].value=occurence[4]
                sh_gin[k]["G"+str(i+2)].value=occurence[5]
                if not occurence[6]==0:
                        sh_gin[k]["H"+str(i+2)].value=occurence[6]
               




#On cherche à présent à attribuer une période aux textes
print("Attribution de la période")
#Liste des fichiers
Fichiers=["neo-assyr.txt","Neo-bab.txt","Achaemenid.txt","Hellenistic.txt","Middle Assyrian.txt","Middle Babylonian.txt", "Old Babylonian.txt", "ED IIIb.txt"]
#Liste des noms d'époques
NomsEpoques=["Neo-Assyrian (ca. 911-612 BC)","Neo-Babylonian (ca. 626-539 BC)","Achaemenid (547-331 BC)","Hellenistic (323-63 BC)","Middle Assyrian (ca. 1400-1000 BC)","Middle Babylonian (ca. 1400-1100 BC)","Old Babylonian (ca. 1900-1600 BC)","ED IIIb (ca. 2500-2340 BC)"]

#On parcourt les différents fichiers. L'idée est qu'on regarde, en parcourant le fichier associé à une époque donnée, les numéros de texte trouvés et on cherche ce numéro dans le fichier Excel.
indice=0
for nom_fichier in Fichiers:
        print(nom_fichier)

        fichier_cdli = open(os.path.join(chemin_data,"Data chronologie",nom_fichier),"r")
        Liste_lignes = fichier_cdli.readlines()
        fichier_cdli.close()





        k=0
        epoque=NomsEpoques[indice]
        marqueur=1
        for ligne in Liste_lignes:
                if k%1000==0:
                        print("*ligne ",k," dans le document txt")#pour s'assurer à execution que le programme fonctionne bien.
                if ligne==[]:
                        continue
                if beginning_recipe(ligne):
                        nom=ligne[1:]
                        for i in range(1,sheet.max_row+1):
                                #if i%100==0:
                                #       print("*      Ligne ",i," dans le fichier XSL")#pour s'assurer que le programme fonctionne.
                                if nom==sheet["A"+str(i)].value:
                                        if not sheet.cell(row=i,column=dd+4).value==None:
                                               m=sheet.cell(row=i,column=dd+4).value
                                               sheet.cell(row=i,column=dd+4).value=m+";"+epoque
                                        else:
                                                sheet.cell(row=i,column=dd+4).value=epoque
                                        marqueur=i
                                        break
                k=k+1
        indice=indice+1
        print(nom_fichier)


#On cherche à présent à attribuer une provenance. On fait comme avec l'époque.
print("Attribution de la période")
#Liste des fichiers
Fichiers_prov=["assur.txt","babylon.txt","Nineveh (Kuyunjik).txt","Nippur.txt","sippar-yahrurum.txt","Uruk.txt"]
Nom_prov=["Assur (mod. Qalat Sherqat)","Bābili (mod. Babylon)","Nineveh (mod. Kuyunjik)","Nippur (mod. Nuffar)","Sippar-Yahrurum (mod. Tell Abu Habbah)","Uruk (mod. Warka)"]

indice=0
for nom_fichier in Fichiers_prov:
        print(nom_fichier)

        fichier_cdli = open(os.path.join(chemin_data,"Data provenance",nom_fichier),"r")
        Liste_lignes = fichier_cdli.readlines()
        fichier_cdli.close()





        k=0
        provenance=Nom_prov[indice]
        marqueur=1
        for ligne in Liste_lignes:
                if k%1000==0:
                        print("*ligne ",k," dans le document txt")
                if ligne==[]:
                        continue
                if beginning_recipe(ligne):
                        nom=ligne[1:]
                        for i in range(1,sheet.max_row+1):
                                #if i%100==0:
                                #       print("*      Ligne ",i," dans le fichier XSL")
                                if nom==sheet["A"+str(i)].value:
                                        if not sheet.cell(row=i,column=dd+5).value==None:
                                               m=sheet.cell(row=i,column=dd+5).value
                                               sheet.cell(row=i,column=dd+5).value=m+";"+provenance
                                        else:
                                                sheet.cell(row=i,column=dd+5).value=provenance
                                        marqueur=i
                                        break
                k=k+1
        indice=indice+1
        print(nom_fichier)


workbook.save(filename="cdli_analysis.xlsx")





